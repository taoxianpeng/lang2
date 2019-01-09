#!usr/bin/python
# -*- coding: utf-8 -*-

# 负责对excel文件的读取与读取
# (读取内容发送给listingword文件，listingword文件反馈的信息，根据单词写入到excel)

from os import getcwd
from os.path import exists
from time import strftime, localtime
from openpyxl import Workbook
from openpyxl import load_workbook


class iosys():
    def __init__(self):
        self.cwd = getcwd()
        self.ntime = strftime("%Y-%m-%d", localtime())
        self.excel_path = ''
        self.list_en = []
        self.list_zh = []

    def setFileName(self,name):
        self.excel_path = name

    def createNew(self):
        if exists(self.excel_path):
            return 'The file had exist!'
        try:
            wb = Workbook()
            sheet = wb.active
            sheet['A1'] = '此列填写单词'
            wb.save(self.excel_path)
            return (201, '保存成功,请打开Excel文件进行编辑')
        except Exception as e:
            print(e)
            return (101, '创建Excel失败!')

    def readExcel_en(self):
        # 读取翻译(tls)、单词(en)
        if not exists(self.excel_path):
            info = self.createNew()
            return info
        try:
            wb = load_workbook(self.excel_path)
            sheet_ranges = wb['Sheet']
            if sheet_ranges['A1'].value == '此列填写单词':
                print('**单词列格式错误，请检查！**')
                return (102, '单词列格式错误，请检查！')
            cell_A = 1
            while True:
                text = sheet_ranges['A{num}'.format(num=cell_A)].value
                if text:
                    self.list_en.append(text)
                    cell_A += 1
                else:
                    return self.list_en
        except Exception as e:
            print(e)
            return (103, e)

    def readExcel_zh(self):
        # 读取翻译(tls)、单词(en)
        if not exists(self.excel_path):
            info = self.createNew()
            return info
        try:

            wb = load_workbook(self.excel_path)
            sheet_ranges = wb['Sheet']
            # if sheet_ranges['B1'].value == None:
            #     print('先翻译单词后再执行操作')
            #     exit()
            cell_B = 1
            while True:
                text = sheet_ranges['B{num}'.format(num=cell_B)].value
                if text:
                    self.list_zh.append(text)
                    cell_B += 1
                else:
                    return self.list_zh

        except Exception as e:
            return (105, e)

    def writeExcel(self, translation):
        zh = translation
        try:
            wb = load_workbook(self.excel_path)
            sheet = wb['Sheet']
            cell_length = len(zh)
            for i in range(cell_length):
                sheet['B{num}'.format(num=i+1)] = zh[i]
            wb.save(self.excel_path)
            print('保存成功')
            return (202, '写入成功，保存成功')
        except Exception as e:
            return (106, e)


if __name__ == '__main__':
    pass
    # one = io()
    # print(one.readExcel())
